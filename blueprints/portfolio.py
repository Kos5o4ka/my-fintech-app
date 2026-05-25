import hashlib
import io
import csv
import json
import logging
import math
import re
from datetime import datetime, date, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import StringIO
from pydantic import ValidationError

from flask import Blueprint, request, jsonify, render_template, abort, make_response
from flask_login import login_required, current_user

from extensions import db, cache
from models import BondPortfolio, Watchlist, Transaction
from moex import (
    get_moex_bond,
    get_bond_history_all,
    search_bonds,
    _fetch_json,
    get_rgbi_history,
    get_screener_bonds,
)
from services.moex_service import (
    get_bond_cached,
    get_bond_preview,
    get_coupon_calendar_cached,
)
from services.portfolio_service import (
    build_portfolio_list,
    calc_portfolio_ytm,
    calc_coupon_income,
    calc_monthly_profit,
    calc_tax_report,
    calc_sharpe_ratio,
)
from schemas.portfolio import AddBondRequest, SellBondRequest, ScreenerRequest
from constants import (
    INCOME_TTL,
    CHART_RANGE_TTL,
    CHART_ALL_TTL,
    BENCHMARK_TTL,
    SCREENER_TTL,
    MAX_CHART_POINTS,
    STATS_TTL,
    DEFAULT_PAGE_SIZE,
    MAX_PAGE_SIZE,
    TIMEFRAME_DAYS,
)

logger = logging.getLogger(__name__)
portfolio_bp = Blueprint("portfolio", __name__)


# ── Утилиты импорта (module-level для тестируемости) ──────────────────────────


def _norm_hdr(v) -> str:
    """Нормализует заголовок: убирает переносы, лишние пробелы, lowercase."""
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").replace("\r", " ").split()).lower()


def _parse_num(v):
    """Парсит число в русском формате: '100,70' / '2 014,00' → float."""
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = (
        str(v)
        .strip()
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _parse_any_date(v):
    """Парсит дату из строки/datetime; понимает DD.MM.YYYY, ISO и другие форматы."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return date.today()
    s = str(v).strip().split()[0].split("/")[0].strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.today()


def _is_valid_isin(s: str) -> bool:
    """ISIN — ровно 12 alphanumeric символов, начинается с 2 букв."""
    return len(s) == 12 and s.isalnum() and s[:2].isalpha()


def _tx_type(val) -> str:
    if not val:
        return "buy"
    v = str(val).strip().lower()
    if v in ("продажа", "sell", "s", "-", "реализация", "погашение"):
        return "sell"
    return "buy"


def _is_repo(val) -> bool:
    """Возвращает True для РЕПО-сделок (не нужны в портфеле)."""
    if not val:
        return False
    return "репо" in str(val).lower() or "repo" in str(val).lower()


def _is_cancelled(val) -> bool:
    if not val:
        return False
    v = str(val).strip().lower()
    return v in ("отменена", "отменено", "cancelled", "canceled", "rejected")


# ── Список псевдонимов столбцов для парсинга брокерских отчётов ───────────────
_ISIN = ["isin", "isin код", "код актива", "код бумаги", "код инструмента", "code"]
_AMT = ["количество", "кол-во", "кол.", "amount", "qty", "объем", "объём"]
_PRICE = ["цена за единицу", "цена сделки", "цена", "price", "курс"]
_DATE = [
    "дата заключения",
    "дата сделки",
    "дата",
    "date",
    "дата операции",
    "дата торгов",
]
_TYPE = ["вид сделки", "тип сделки", "операция", "тип операции", "направление", "type"]
_NAME = [
    "наименование актива",
    "наименование инструмента",
    "наименование",
    "название",
    "инструмент",
    "name",
]
_COMM = [
    "комиссия брокера",
    "сумма комиссии брокера",
    "комиссия",
    "commission",
    "broker_commission",
]
_CURR = ["валюта расчетов", "валюта расчётов", "валюта", "currency"]
_STATUS = ["признак исполнения", "статус", "status"]
_ANCHORS = _ISIN + _AMT + _PRICE
_PRICE_CURR = ["валюта цены", "единица цены", "валюта цены сделки"]
_DEAL_NO = ["номер сделки", "№ сделки", "n сделки", "номер"]


def _find_header_row_xlsx(sheet, max_scan: int = 50):
    """Возвращает (row_idx, {norm_name: col_idx}) или (None, {}).
    Работает с обычным (non-read_only) режимом openpyxl.
    """
    anchors = {_norm_hdr(a) for a in _ANCHORS}
    max_r = sheet.max_row or max_scan
    for ri in range(1, min(max_scan + 1, max_r + 1)):
        hdrs = {}
        for cell in sheet[ri]:
            n = _norm_hdr(cell.value)
            if n:
                hdrs[n] = cell.column
        if sum(1 for a in anchors if a in hdrs) >= 2:
            return ri, hdrs
    return None, {}


def _parse_vtb_xlsx(all_rows: list) -> list:
    """Парсер брокерского отчёта ВТБ (.xlsx).

    Разделы «Заключённые / Завершённые в отчёте сделки»:
      col 1  — «Наименование, Рег.код, ISIN»
      col 2  — дата и время заключения (datetime)
      col 5  — вид сделки (Покупка / Продажа)
      col 7  — количество (шт.)
      col 9  — цена (% от номинала для облигаций)
      col 11 — валюта расчётов (RUR → нормализуем в RUB)
      col 15 — комиссия Банка за расчёты
      col 17 — комиссия Банка за заключение
      col 25 — № сделки (дедупликация между разделами)

    Фильтр облигаций: рег.код начинается на «4B» (корп./биржевые)
    или содержит «RMFS» (ОФЗ).
    """
    seen_deals: set = set()
    deals: list = []

    for row in all_rows:
        if not row or len(row) < 18:
            continue

        type_cell = row[5]
        if not type_cell:
            continue
        type_s = str(type_cell).strip().lower()
        if type_s not in ("покупка", "продажа"):
            continue

        name_cell = row[1]
        if not name_cell:
            continue

        parts = [p.strip() for p in str(name_cell).split(",")]
        if len(parts) < 3:
            continue

        isin = parts[-1].upper()
        if not _is_valid_isin(isin):
            continue

        reg_code = parts[-2].upper()
        if not (reg_code.startswith("4B") or "RMFS" in reg_code):
            continue

        deal_no = (
            str(row[25]).strip()
            if len(row) > 25 and row[25] is not None
            else ""
        )
        if deal_no and deal_no in seen_deals:
            continue
        if deal_no:
            seen_deals.add(deal_no)

        price_pct = row[9]
        qty_v = row[7]
        date_v = row[2]
        comm1 = row[15] if len(row) > 15 else None
        comm2 = row[17] if len(row) > 17 else None
        curr_v = row[11] if len(row) > 11 else None

        try:
            price_rub = float(price_pct) * 10
        except (TypeError, ValueError):
            continue
        if price_rub <= 0:
            continue

        commission = 0.0
        for c in (comm1, comm2):
            try:
                commission += float(c)
            except (TypeError, ValueError):
                pass

        raw_curr = str(curr_v or "").strip().upper()
        if raw_curr in ("RUR", "RUB"):
            currency = "RUB"
        elif raw_curr.isalpha() and len(raw_curr) == 3:
            currency = raw_curr
        else:
            currency = "RUB"

        deals.append(
            {
                "isin": isin,
                "amount": qty_v,
                "price": price_rub,
                "date": date_v,
                "tx_type": "sell" if type_s == "продажа" else "buy",
                "name": ", ".join(parts[:-2]),
                "commission": commission if commission > 0 else None,
                "currency": currency,
                "notes": None,
            }
        )

    return deals


def _detect_broker(all_rows: list) -> str:
    """Определяет брокера по сигнатурам XLSX-отчёта.

    Возвращает 'vtb', 'tinkoff' или 'generic'.
    """
    # ВТБ: строки данных с col[5] = Покупка/Продажа и col[1] = "Назв, РегКод, ISIN"
    vtb_votes = 0
    for row in all_rows[:100]:
        if not row or len(row) < 6:
            continue
        if str(row[5] or "").strip().lower() in ("покупка", "продажа"):
            if str(row[1] or "").count(",") >= 2:
                vtb_votes += 1
                if vtb_votes >= 2:
                    return "vtb"

    # Т-Инвестиции: ищем маркеры в первых строках метаданных
    tinkoff_markers = ("т-инвестиции", "tinkoff", "tbank", "т инвестиции")
    for row in all_rows[:40]:
        if not row:
            continue
        for cell in row:
            if not cell:
                continue
            s = str(cell).strip().lower()
            if any(m in s for m in tinkoff_markers):
                return "tinkoff"

    return "generic"


def _find_header_row_from_list(rows: list, max_scan: int = 50):
    """Возвращает (row_idx_0based, {norm_name: col_idx_1based}) или (None, {}).
    Работает с результатом iter_rows(values_only=True) — списком кортежей.
    """
    anchors = {_norm_hdr(a) for a in _ANCHORS}
    for ri, row_values in enumerate(rows[:max_scan]):
        hdrs = {}
        for col_idx, cell_val in enumerate(row_values, start=1):
            n = _norm_hdr(cell_val)
            if n:
                hdrs[n] = col_idx
        if sum(1 for a in anchors if a in hdrs) >= 2:
            return ri, hdrs
    return None, {}


def _find_col(hdrs: dict, candidates: list):
    """Первый совпадающий индекс столбца (1-based) или None."""
    for name in candidates:
        n = _norm_hdr(name)
        if n in hdrs:
            return hdrs[n]
    for name in candidates:
        n = _norm_hdr(name)
        for hn, hc in hdrs.items():
            if hn.startswith(n) or n.startswith(hn):
                return hc
    return None


# ── Прочие утилиты blueprint'а ─────────────────────────────────────────────────


def _etag(payload: dict) -> str:
    """Быстрый ETag из MD5 тела ответа (первые 16 символов)."""
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _bust_user_cache(user_id: int) -> None:
    """Инвалидирует кэши, зависящие от состава портфеля пользователя."""
    for key in [
        f"portfolio_income:{user_id}",
        f"portfolio_stats:{user_id}",
    ]:
        try:
            cache.delete(key)
        except Exception:
            pass


# ── Страница ──────────────────────────────────────────────────────────────────


@portfolio_bp.route("/portfolio")
@login_required
def portfolio_page() -> str:
    return render_template("portfolio.html")


# ── Активный портфель ─────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio", methods=["GET"])
@login_required
def get_portfolio():
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = min(
        request.args.get("per_page", DEFAULT_PAGE_SIZE, type=int), MAX_PAGE_SIZE
    )

    q = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False)
    total_count = q.count()
    active = (
        q.order_by(BondPortfolio.id).offset((page - 1) * per_page).limit(per_page).all()
    )

    bonds, total_val = build_portfolio_list(active)
    ytm = calc_portfolio_ytm(bonds, total_val)

    payload = {
        "status": "success",
        "total_value": total_val,
        "portfolio_ytm": ytm,
        "bonds": bonds,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total_count,
            "pages": math.ceil(total_count / per_page) if per_page else 1,
        },
    }

    # ETag — клиент может прислать If-None-Match и получить 304
    tag = _etag(payload)
    if request.headers.get("If-None-Match") == tag:
        return "", 304

    resp = make_response(jsonify(payload))
    resp.headers["ETag"] = tag
    resp.headers["Cache-Control"] = "private, max-age=60"
    return resp


# ── Заметки к позиции ────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/<int:bond_id>/notes", methods=["PATCH"])
@login_required
def update_bond_notes(bond_id: int):
    """Обновляет заметку к позиции портфеля (Stage 4)."""
    bond = BondPortfolio.query.filter_by(
        id=bond_id, user_id=current_user.id
    ).first_or_404()
    data = request.get_json() or {}
    raw = (data.get("notes") or "").strip()
    bond.notes = raw if raw else None
    db.session.commit()
    return jsonify({"status": "success", "notes": bond.notes or ""})


def build_transaction_entry(t: Transaction) -> dict:
    buy_p = float(t.price)
    sell_p = float(t.price)
    commission = float(t.commission) if t.commission else 0.0
    pnl = 0.0
    pnl_pct = 0.0
    sell_date = None
    purchase_date = None

    if t.tx_type == "sell":
        sell_date = t.tx_date.strftime("%Y-%m-%d")
        # Try to find corresponding BondPortfolio record to get buy_price
        bond = BondPortfolio.query.filter_by(
            user_id=t.user_id,
            isin=t.isin,
            is_sold=True,
            amount=t.amount,
            sell_date=t.tx_date,
        ).first()
        if bond:
            buy_p = float(bond.buy_price)
            sell_p = float(bond.sell_price) if bond.sell_price else float(t.price)
            commission = (
                float(bond.broker_commission) if bond.broker_commission else commission
            )
            pnl = (sell_p - buy_p) * t.amount - commission
            pnl_pct = (pnl / (buy_p * t.amount) * 100) if buy_p else 0.0
            if bond.purchase_date:
                purchase_date = bond.purchase_date.strftime("%Y-%m-%d")
    else:
        purchase_date = t.tx_date.strftime("%Y-%m-%d")

    return {
        "id": t.id,
        "isin": t.isin,
        "name": t.name or "Облигация",
        "tx_type": t.tx_type,
        "amount": t.amount,
        "buy_price": buy_p,
        "sell_price": round(sell_p, 2),
        "commission": round(commission, 2),
        "pnl": round(pnl, 2),
        "pnl_pct": round(pnl_pct, 2),
        "purchase_date": purchase_date,
        "sell_date": sell_date,
        "date": t.tx_date.strftime("%Y-%m-%d"),
        "currency": t.currency or "RUB",
    }


# ── История сделок ────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/history", methods=["GET"])
@login_required
def portfolio_history():
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = min(
        request.args.get("per_page", DEFAULT_PAGE_SIZE, type=int), MAX_PAGE_SIZE
    )
    date_from_str = request.args.get("date_from", "").strip()
    date_to_str = request.args.get("date_to", "").strip()
    tx_type = request.args.get("tx_type")

    # Dynamic self-healing synthesis from BondPortfolio if Transactions are empty
    tx_count = Transaction.query.filter_by(user_id=current_user.id).count()
    if tx_count == 0:
        bonds = BondPortfolio.query.filter_by(user_id=current_user.id).all()
        for b in bonds:
            buy_tx = Transaction(
                user_id=current_user.id,
                isin=b.isin,
                name=b.name,
                tx_type="buy",
                amount=b.amount,
                price=b.buy_price,
                tx_date=b.purchase_date,
                commission=0.0,
            )
            db.session.add(buy_tx)
            if b.is_sold:
                sell_tx = Transaction(
                    user_id=current_user.id,
                    isin=b.isin,
                    name=b.name,
                    tx_type="sell",
                    amount=b.amount,
                    price=b.sell_price if b.sell_price is not None else b.buy_price,
                    tx_date=b.sell_date if b.sell_date is not None else b.purchase_date,
                    commission=b.broker_commission,
                )
                db.session.add(sell_tx)
        db.session.commit()

    if tx_type is None:
        tx_type = "sell"

    query = Transaction.query.filter_by(user_id=current_user.id)
    if tx_type in ("buy", "sell"):
        query = query.filter_by(tx_type=tx_type)

    if date_from_str:
        try:
            query = query.filter(
                Transaction.tx_date
                >= datetime.strptime(date_from_str, "%Y-%m-%d").date()
            )
        except ValueError:
            pass
    if date_to_str:
        try:
            query = query.filter(
                Transaction.tx_date <= datetime.strptime(date_to_str, "%Y-%m-%d").date()
            )
        except ValueError:
            pass

    total_count = query.count()
    tx_list = (
        query.order_by(Transaction.tx_date.desc(), Transaction.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return jsonify(
        {
            "status": "success",
            "trades": [build_transaction_entry(t) for t in tx_list],
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total_count,
                "pages": math.ceil(total_count / per_page) if per_page else 1,
            },
        }
    )


# ── Поиск облигаций ───────────────────────────────────────────────────────────


@portfolio_bp.route("/api/search_bond", methods=["GET"])
@login_required
def search_bond():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify(search_bonds(q, limit=8))


# ── Превью облигации ──────────────────────────────────────────────────────────


@portfolio_bp.route("/api/bond_preview/<isin>", methods=["GET"])
@login_required
def bond_preview(isin: str):
    isin = isin.upper().strip()
    result = get_bond_preview(isin)
    if not result:
        return jsonify(
            {"status": "error", "message": "Облигация не найдена на Московской Бирже"}
        ), 404
    return jsonify(result)


# ── Добавить облигацию ────────────────────────────────────────────────────────


@portfolio_bp.route("/api/add_bond", methods=["POST"])
@login_required
def add_bond():
    try:
        req = AddBondRequest.model_validate(request.get_json() or {})
    except ValidationError as e:
        first_error = e.errors()[0]["msg"].replace("Value error, ", "")
        return jsonify({"status": "error", "message": first_error}), 400

    isin = req.isin
    purchase_date = (
        datetime.strptime(req.purchase_date, "%Y-%m-%d").date()
        if req.purchase_date
        else date.today()
    )

    moex_data = get_moex_bond(isin)
    if not moex_data:
        return jsonify(
            {
                "status": "error",
                "message": f"Облигация {isin} не найдена на Московской Бирже.",
            }
        ), 404

    secid = moex_data["secid"]
    bond_title = moex_data.get("name", "Облигация")

    # Проверка дат размещения и погашения
    try:
        res = _fetch_json(f"https://iss.moex.com/iss/securities/{secid}.json")
        if res.get("description") and res["description"].get("data"):
            issue_date: Optional[date] = None
            mat_date: Optional[date] = None
            for row in res["description"]["data"]:
                if row[0] == "ISSUEDATE" and row[2]:
                    issue_date = datetime.strptime(row[2], "%Y-%m-%d").date()
                if row[0] == "MATDATE" and row[2]:
                    mat_date = datetime.strptime(row[2], "%Y-%m-%d").date()
            if issue_date and purchase_date < issue_date:
                return jsonify(
                    {
                        "status": "error",
                        "message": f"Ошибка валидации: облигация выпущена {issue_date}. Нельзя купить бумагу до эмиссии.",
                    }
                ), 400
            if mat_date and purchase_date > mat_date:
                return jsonify(
                    {
                        "status": "error",
                        "message": f"Ошибка валидации: облигация погашена {mat_date}. Торги закрыты.",
                    }
                ), 400
    except Exception as exc:
        logger.warning("Date spec validation error for %s: %s", secid, exc)

    existing = BondPortfolio.query.filter_by(
        user_id=current_user.id, isin=isin, is_sold=False
    ).first()
    existing_amount = existing.amount if existing else 0

    live_price = moex_data.get("price", float(req.buy_price))
    currency = moex_data.get("currency", "RUB")
    new_bond = BondPortfolio(
        user_id=current_user.id,
        isin=isin,
        secid=secid,
        name=bond_title,
        amount=int(req.amount),
        buy_price=float(req.buy_price),
        last_price=live_price,
        purchase_date=purchase_date,
        is_sold=False,
        currency=currency,
        notes=req.notes.strip() if req.notes else None,
    )
    db.session.add(new_bond)
    db.session.add(
        Transaction(
            user_id=current_user.id,
            isin=isin,
            name=bond_title,
            tx_type="buy",
            amount=int(req.amount),
            price=float(req.buy_price),
            currency=currency,
            tx_date=purchase_date,
        )
    )
    db.session.commit()
    _bust_user_cache(current_user.id)
    return jsonify(
        {
            "status": "success",
            "message": f"Бумага {bond_title} успешно добавлена!",
            "duplicate_warning": existing_amount > 0,
            "existing_amount": existing_amount,
        }
    ), 201


# ── Продать облигацию ─────────────────────────────────────────────────────────


@portfolio_bp.route("/api/sell_bond/<int:bond_id>", methods=["POST"])
@login_required
def sell_bond(bond_id: int):
    bond = db.session.get(BondPortfolio, bond_id)
    if bond is None:
        abort(404)
    if bond.user_id != current_user.id:
        abort(403)

    try:
        req = SellBondRequest.model_validate(request.get_json() or {})
    except ValidationError as e:
        first_error = e.errors()[0]["msg"].replace("Value error, ", "")
        return jsonify({"status": "error", "message": first_error}), 400

    if req.amount and req.amount > bond.amount:
        return jsonify(
            {
                "status": "error",
                "message": f"Нельзя продать больше, чем есть в наличии ({bond.amount} шт.).",
            }
        ), 400

    sell_price = (
        req.sell_price
        if req.sell_price
        else (float(bond.last_price) if bond.last_price else float(bond.buy_price))
    )

    if req.amount and req.amount < bond.amount:
        sell_qty = req.amount
        bond.amount -= sell_qty
        sold_bond = BondPortfolio(
            user_id=bond.user_id,
            isin=bond.isin,
            secid=bond.secid,
            name=bond.name,
            amount=sell_qty,
            buy_price=bond.buy_price,
            last_price=bond.last_price,
            purchase_date=bond.purchase_date,
            is_sold=True,
            sell_date=date.today(),
            sell_price=sell_price,
            broker_commission=req.broker_commission,
            notes=bond.notes,
        )
        db.session.add(sold_bond)
        message = f"Частично продано {sell_qty} шт. облигации {bond.name}."
    else:
        sell_qty = bond.amount
        bond.is_sold = True
        bond.sell_date = date.today()
        bond.sell_price = sell_price
        if req.broker_commission is not None:
            bond.broker_commission = req.broker_commission
        sold_bond = bond
        message = f"Облигация {bond.name} полностью продана и переведена в архив."

    db.session.add(
        Transaction(
            user_id=current_user.id,
            isin=bond.isin,
            name=bond.name,
            tx_type="sell",
            amount=sell_qty,
            price=sell_price,
            commission=req.broker_commission,
            tx_date=sold_bond.sell_date,
            currency=bond.currency or "RUB",
        )
    )
    db.session.commit()
    _bust_user_cache(current_user.id)
    return jsonify({"status": "success", "message": message})


# ── График цены облигации ─────────────────────────────────────────────────────


@portfolio_bp.route("/api/bond_chart/<isin>", methods=["GET"])
@login_required
def get_bond_chart_data(isin: str):
    range_param = request.args.get("range", "all")
    cache_key = f"bond_chart:{isin}:{range_param}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)

    moex_data = get_moex_bond(isin)
    if not moex_data:
        return jsonify({"status": "error", "message": "Бумага не найдена"}), 404

    full = get_bond_history_all(moex_data["secid"], moex_data.get("facevalue", 1000))
    labels = full.get("labels", [])
    prices = full.get("data", [])
    nkd_hist = full.get("nkd", [])
    ytm_hist = full.get("ytm", [])

    if range_param in ("day", "week", "month"):
        cutoff = datetime.utcnow().date() - timedelta(days=TIMEFRAME_DAYS[range_param])
        combined = [
            (lbl, p, n, y)
            for lbl, p, n, y in zip(labels, prices, nkd_hist, ytm_hist)
            if _parse_date(lbl) and _parse_date(lbl) >= cutoff
        ]
        if not combined:
            take = min(100, len(labels))
            combined = list(
                zip(labels[-take:], prices[-take:], nkd_hist[-take:], ytm_hist[-take:])
            )
    else:
        combined = list(zip(labels, prices, nkd_hist, ytm_hist))

    if len(combined) > MAX_CHART_POINTS:
        step = math.ceil(len(combined) / MAX_CHART_POINTS)
        combined = [combined[i] for i in range(0, len(combined), step)]

    if combined:
        labels_out, prices_out, nkd_out, ytm_out = zip(*combined)
        result = {
            "labels": list(labels_out),
            "data": list(prices_out),
            "nkd": list(nkd_out),
            "ytm": list(ytm_out),
        }
    else:
        result = {"labels": [], "data": [], "nkd": [], "ytm": []}

    ttl = CHART_RANGE_TTL if range_param != "all" else CHART_ALL_TTL
    try:
        cache.set(cache_key, result, timeout=ttl)
    except Exception:
        pass
    return jsonify(result)


def _parse_date(lbl: str) -> date | None:
    try:
        return datetime.strptime(lbl, "%Y-%m-%d").date()
    except Exception:
        return None


# ── Купонный календарь ────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/calendar", methods=["GET"])
@login_required
def get_portfolio_calendar():
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    events = []
    for bond in active:
        target = bond.secid or bond.isin
        for c in get_coupon_calendar_cached(target):
            events.append(
                {
                    "name": bond.name or bond.isin,
                    "isin": bond.isin,
                    "date": c["date"],
                    "total_payout": round(c["value"] * bond.amount, 2),
                }
            )
    events.sort(key=lambda x: x["date"])
    return jsonify(events[:10])


# ── Прогноз купонного дохода ──────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/income", methods=["GET"])
@login_required
def portfolio_income():
    cache_key = f"portfolio_income:{current_user.id}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    result = calc_coupon_income(active)
    cache.set(cache_key, result, timeout=INCOME_TTL)
    return jsonify(result)


# ── Распределение по эмитентам ────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/allocation", methods=["GET"])
@login_required
def portfolio_allocation():
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    slices = [
        {
            "name": b.name or b.isin,
            "value": round(
                (float(b.last_price) if b.last_price else float(b.buy_price))
                * b.amount,
                2,
            ),
        }
        for b in active
    ]
    slices.sort(key=lambda x: x["value"], reverse=True)
    return jsonify(slices)


# ── Экспорт CSV ───────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/export", methods=["GET"])
@login_required
def export_portfolio_csv():
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(
        [
            "Название бумаги",
            "ISIN код",
            "Количество (шт)",
            "Цена покупки (руб)",
            "Дата сделки",
        ]
    )
    for bond in active:
        cw.writerow(
            [bond.name, bond.isin, bond.amount, bond.buy_price, bond.purchase_date]
        )
    response = make_response("﻿" + si.getvalue())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.csv"
    )
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    return response


# ── Экспорт XLSX ──────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/export/xlsx", methods=["GET"])
@login_required
def export_portfolio_xlsx():
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    sold = (
        BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True)
        .order_by(BondPortfolio.sell_date.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "Портфель"
    headers1 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Посл. цена (₽)",
        "P&L (₽)",
        "Дата покупки",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1E7E34")
        cell.alignment = center
    for bond in active:
        last_p = float(bond.last_price) if bond.last_price else float(bond.buy_price)
        pnl = round((last_p - float(bond.buy_price)) * bond.amount, 2)
        ws1.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                float(bond.buy_price),
                round(last_p, 2),
                pnl,
                bond.purchase_date.strftime("%Y-%m-%d") if bond.purchase_date else "",
            ]
        )
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    ws2 = wb.create_sheet("История сделок")
    headers2 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Цена продажи (₽)",
        "Комиссия (₽)",
        "P&L (₽)",
        "P&L %",
        "Дата продажи",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="155724")
        cell.alignment = center
    for bond in sold:
        buy_p = float(bond.buy_price)
        sell_p = float(bond.sell_price) if bond.sell_price else buy_p
        comm = float(bond.broker_commission) if bond.broker_commission else 0.0
        pnl = round((sell_p - buy_p) * bond.amount - comm, 2)
        pnl_pct = round(pnl / (buy_p * bond.amount) * 100, 2) if buy_p else 0.0
        ws2.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                round(buy_p, 2),
                round(sell_p, 2),
                round(comm, 2),
                pnl,
                pnl_pct,
                bond.sell_date.strftime("%Y-%m-%d") if bond.sell_date else "",
            ]
        )
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.xlsx"
    )
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response


# ── Вотчлист ──────────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/watchlist", methods=["GET"])
@login_required
def get_watchlist():
    items = (
        Watchlist.query.filter_by(user_id=current_user.id)
        .order_by(Watchlist.added_at.desc())
        .all()
    )
    result = []
    for item in items:
        moex_data = get_bond_cached(item.isin) or {}
        result.append(
            {
                "isin": item.isin,
                "name": item.name or item.isin,
                "added_at": item.added_at.strftime("%Y-%m-%d"),
                "price": moex_data.get("price"),
                "ytm": moex_data.get("ytm"),
                "nkd": moex_data.get("nkd"),
            }
        )
    return jsonify(result)


@portfolio_bp.route("/api/watchlist", methods=["POST"])
@login_required
def add_to_watchlist():
    data = request.get_json() or {}
    isin = data.get("isin", "").upper().strip()
    if not isin:
        return jsonify({"status": "error", "message": "ISIN обязателен."}), 400
    if Watchlist.query.filter_by(user_id=current_user.id, isin=isin).first():
        return jsonify(
            {"status": "error", "message": "Облигация уже в избранном."}
        ), 409
    moex_data = get_moex_bond(isin)
    if not moex_data:
        return jsonify(
            {"status": "error", "message": f"Облигация {isin} не найдена на MOEX."}
        ), 404
    entry = Watchlist(
        user_id=current_user.id,
        isin=isin,
        secid=moex_data.get("secid"),
        name=moex_data.get("name", isin),
    )
    db.session.add(entry)
    db.session.commit()
    return jsonify(
        {"status": "success", "message": f"{entry.name} добавлена в избранное."}
    ), 201


@portfolio_bp.route("/api/watchlist/<isin>", methods=["DELETE"])
@login_required
def remove_from_watchlist(isin: str):
    entry = Watchlist.query.filter_by(
        user_id=current_user.id, isin=isin.upper()
    ).first()
    if not entry:
        return jsonify({"status": "error", "message": "Не найдено."}), 404
    db.session.delete(entry)
    db.session.commit()
    return jsonify({"status": "success", "message": "Удалено из избранного."})


# ── Скринер ───────────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/screener", methods=["GET"])
@login_required
def screener():
    try:
        req = ScreenerRequest(
            min_ytm=request.args.get("min_ytm") or None,
            max_ytm=request.args.get("max_ytm") or None,
            maturity_from=request.args.get("maturity_from") or None,
            maturity_to=request.args.get("maturity_to") or None,
            issuer_type=request.args.get("issuer_type") or None,
            min_duration=request.args.get("min_duration") or None,
            max_duration=request.args.get("max_duration") or None,
        )
    except ValidationError:
        return jsonify([])

    cache_key = (
        f"screener:{req.min_ytm}:{req.max_ytm}:{req.maturity_from}:{req.maturity_to}"
    )
    cached = cache.get(cache_key)
    if not cached:
        cached = get_screener_bonds(
            req.min_ytm, req.max_ytm, req.maturity_from, req.maturity_to, limit=200
        )
        cache.set(cache_key, cached, timeout=SCREENER_TTL)

    results = cached

    # Stage 4: post-filter by issuer type (client-side semantic)
    if req.issuer_type:
        itype = req.issuer_type.lower()

        def _matches_type(b: dict) -> bool:
            name = (b.get("name") or b.get("secid") or "").upper()
            if itype == "ofz":
                return (
                    "ОФЗ" in name or name.startswith("SU") or name.startswith("RU000A0")
                )
            if itype == "muni":
                return any(
                    k in name
                    for k in (
                        "МУН",
                        "МУНИЦИПАЛ",
                        "ОБЛИГАЦ",
                        "РЕГИОН",
                        "ОБЛАСТЬ",
                        "КРАЙ",
                        "ГОРОД",
                    )
                )
            if itype == "corp":
                return "ОФЗ" not in name and not any(
                    k in name for k in ("МУН", "РЕГИОН", "ОБЛАСТЬ", "КРАЙ", "ГОРОД")
                )
            return True

        results = [b for b in results if _matches_type(b)]

    # Stage 4: post-filter by approximate duration (years to maturity)
    if req.min_duration is not None or req.max_duration is not None:
        today_ord = date.today().toordinal()
        filtered = []
        for b in results:
            mat = b.get("maturity_date") or b.get("matdate") or ""
            if not mat:
                filtered.append(b)
                continue
            try:
                dur = (date.fromisoformat(mat[:10]).toordinal() - today_ord) / 365.25
            except ValueError:
                filtered.append(b)
                continue
            if req.min_duration is not None and dur < req.min_duration:
                continue
            if req.max_duration is not None and dur > req.max_duration:
                continue
            filtered.append(b)
        results = filtered

    return jsonify(results)


# ── Налоговый отчёт ───────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/tax", methods=["GET"])
@login_required
def portfolio_tax():
    """Налоговый отчёт за год — сделки + купоны + НДФЛ (Stage 4 UI)."""
    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    sold = BondPortfolio.query.filter(
        BondPortfolio.user_id == current_user.id,
        BondPortfolio.is_sold == True,  # noqa: E712
        BondPortfolio.sell_date >= year_start,
        BondPortfolio.sell_date <= year_end,
    ).all()
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    summary = calc_tax_report(sold, active, year)

    # Build enriched response for Stage 4 Tax UI
    trades_list = []
    gross_profit = 0.0
    total_commission = 0.0
    for bond in sold:
        buy_p = float(bond.buy_price)
        sell_p = float(bond.sell_price) if bond.sell_price else buy_p
        comm = float(bond.broker_commission) if bond.broker_commission else 0.0
        pnl = (sell_p - buy_p) * bond.amount - comm
        gross_profit += pnl
        total_commission += comm
        trades_list.append(
            {
                "id": bond.id,
                "name": bond.name or bond.isin,
                "isin": bond.isin,
                "amount": bond.amount,
                "buy_price": round(buy_p, 2),
                "sell_price": round(sell_p, 2),
                "commission": round(comm, 2),
                "pnl": round(pnl, 2),
                "sell_date": bond.sell_date.strftime("%Y-%m-%d")
                if bond.sell_date
                else None,
            }
        )

    taxable_base = max(0.0, round(gross_profit, 2))
    tax_amount = round(taxable_base * 0.13, 2)

    return jsonify(
        {
            **summary,
            "gross_profit": round(gross_profit, 2),
            "total_commission": round(total_commission, 2),
            "taxable_base": taxable_base,
            "tax_amount": tax_amount,
            "trades": trades_list,
        }
    )


# ── Бенчмарк RGBI ─────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/benchmark", methods=["GET"])
@login_required
def portfolio_benchmark():
    range_param = request.args.get("range", "month")
    cache_key = f"benchmark:{range_param}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)

    days = TIMEFRAME_DAYS.get(range_param, 31)
    from_date = (
        (date.today() - timedelta(days=days)).strftime("%Y-%m-%d")
        if days < 9999
        else None
    )
    rgbi = get_rgbi_history(
        from_date=from_date, to_date=date.today().strftime("%Y-%m-%d")
    )
    result = {"range": range_param, "rgbi": rgbi}
    cache.set(cache_key, result, timeout=BENCHMARK_TTL)
    return jsonify(result)


# ── Sharpe Ratio ─────────────────────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio/sharpe", methods=["GET"])
@login_required
def portfolio_sharpe():
    """Коэффициент Шарпа по закрытым позициям портфеля (Stage 4)."""
    sold = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True).all()
    result = calc_sharpe_ratio(sold)
    if result is None:
        return jsonify(
            {
                "sharpe": None,
                "reason": f"Недостаточно данных (закрытых позиций: {len(sold)}, нужно ≥ 3)",
            }
        )
    return jsonify(result)


# ── Сравнение двух облигаций ──────────────────────────────────────────────────


def _bond_history_for_compare(isin: str, range_param: str) -> dict:
    """Вспомогательная функция: история цены одной облигации для вкладки Сравнение."""
    cache_key = f"bond_chart:{isin}:{range_param}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    moex_data = get_moex_bond(isin)
    if not moex_data:
        return {"labels": [], "data": [], "name": isin}

    full = get_bond_history_all(moex_data["secid"], moex_data.get("facevalue", 1000))
    labels = full.get("labels", [])
    prices = full.get("data", [])

    if range_param in ("day", "week", "month"):
        cutoff = datetime.utcnow().date() - timedelta(days=TIMEFRAME_DAYS[range_param])
        combined = [
            (lbl, p)
            for lbl, p in zip(labels, prices)
            if _parse_date(lbl) and _parse_date(lbl) >= cutoff
        ]
        if not combined:
            take = min(100, len(labels))
            combined = list(zip(labels[-take:], prices[-take:]))
    else:
        combined = list(zip(labels, prices))

    if len(combined) > MAX_CHART_POINTS:
        step = math.ceil(len(combined) / MAX_CHART_POINTS)
        combined = [combined[i] for i in range(0, len(combined), step)]

    if combined:
        lbl_out, price_out = zip(*combined)
        result = {
            "labels": list(lbl_out),
            "data": list(price_out),
            "name": moex_data.get("name", isin),
        }
    else:
        result = {"labels": [], "data": [], "name": moex_data.get("name", isin)}

    cache.set(cache_key, result, timeout=CHART_RANGE_TTL)
    return result


@portfolio_bp.route("/api/portfolio/compare", methods=["GET"])
@login_required
def compare_bonds():
    """Сравнение динамики цен двух облигаций (нормировано к 100).

    ?isin1=RU000A…&isin2=RU000A…&range=month|week|3m|year|all
    """
    isin1 = request.args.get("isin1", "").upper().strip()
    isin2 = request.args.get("isin2", "").upper().strip()
    range_param = request.args.get("range", "month")

    if not isin1 or not isin2:
        return jsonify({"status": "error", "message": "Оба ISIN обязательны"}), 400
    if isin1 == isin2:
        return jsonify({"status": "error", "message": "ISIN должны быть разными"}), 400

    cache_key = f"compare:{isin1}:{isin2}:{range_param}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)

    d1 = _bond_history_for_compare(isin1, range_param)
    d2 = _bond_history_for_compare(isin2, range_param)

    def _normalize(prices: list) -> list:
        if not prices:
            return prices
        base = prices[0]
        return [round(p / base * 100, 2) if base else p for p in prices]

    response = {
        "status": "success",
        "labels": d1["labels"] or d2["labels"],
        "bond1": {"isin": isin1, "name": d1["name"], "data": _normalize(d1["data"])},
        "bond2": {"isin": isin2, "name": d2["name"], "data": _normalize(d2["data"])},
    }
    cache.set(cache_key, response, timeout=CHART_RANGE_TTL)
    return jsonify(response)


# ── Отчёт (print-to-PDF) ──────────────────────────────────────────────────────


@portfolio_bp.route("/portfolio/report")
@login_required
def portfolio_report_page():
    """Страница отчёта, оптимизированная для печати / сохранения как PDF."""
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    bonds, total_val = build_portfolio_list(active)
    ytm = calc_portfolio_ytm(bonds, total_val)

    year = date.today().year
    sold_all = BondPortfolio.query.filter_by(
        user_id=current_user.id, is_sold=True
    ).all()
    sold_year = [b for b in sold_all if b.sell_date and b.sell_date.year == year]
    tax = calc_tax_report(sold_year, active, year)

    sharpe_data = calc_sharpe_ratio(sold_all)

    return render_template(
        "pdf_report.html",
        bonds=bonds,
        total_value=round(total_val, 2),
        portfolio_ytm=ytm,
        tax=tax,
        sharpe=sharpe_data,
        year=year,
        username=current_user.username,
        generated_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
        bond_count=len(active),
        sold_count=len(sold_all),
    )


# ── Dashboard: P&L chart data ─────────────────────────────────────────────────


@portfolio_bp.route("/api/dashboard/pnl_chart", methods=["GET"])
@login_required
def dashboard_pnl_chart():
    """Возвращает данные для area-чарта P&L на дашборде.

    ?period=7d   — последние 7 дней (дневные точки)
    ?period=30d  — последние 30 дней (дневные точки, по 5 подписей)
    ?period=ytd  — с начала года (месячные точки)
    """
    period = request.args.get("period", "30d")
    today = date.today()

    if period == "7d":
        start = today - timedelta(days=6)
        date_range = [start + timedelta(days=i) for i in range(7)]
        label_fmt = "%d.%m"
        # Показываем каждую дату
        tick_every = 1
    elif period == "ytd":
        # Месячные точки с 1 января до сегодня
        start = date(today.year, 1, 1)
        date_range = []
        m = start
        while m <= today:
            date_range.append(m)
            if m.month == 12:
                m = date(m.year + 1, 1, 1)
            else:
                m = date(m.year, m.month + 1, 1)
        label_fmt = "%b"
        tick_every = 1
    else:  # 30d
        start = today - timedelta(days=29)
        date_range = [start + timedelta(days=i) for i in range(30)]
        label_fmt = "%d.%m"
        tick_every = 5  # подпись каждые 5 дней

    # Собираем P&L по дням из закрытых позиций (только DB, без MOEX)
    sold = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True).all()
    daily_pnl: dict[date, float] = {}
    for bond in sold:
        if not bond.sell_date:
            continue
        sell_p = float(bond.sell_price) if bond.sell_price else float(bond.buy_price)
        commission = float(bond.broker_commission) if bond.broker_commission else 0.0
        pnl = (sell_p - float(bond.buy_price)) * bond.amount - commission
        daily_pnl[bond.sell_date] = daily_pnl.get(bond.sell_date, 0.0) + pnl

    # Строим кумулятивный P&L внутри периода (старт = 0 в начале периода)
    if period == "ytd":
        # Накапливаем помесячно
        labels, data = [], []
        running = 0.0
        for d in date_range:
            for sell_date, pnl in daily_pnl.items():
                if sell_date.year == d.year and sell_date.month == d.month:
                    running += pnl
            labels.append(d.strftime(label_fmt))
            data.append(round(running, 2))
    else:
        labels, data = [], []
        running = 0.0
        for i, d in enumerate(date_range):
            running += daily_pnl.get(d, 0.0)
            # Показываем подпись не для каждой точки — chart.js разберётся
            labels.append(
                d.strftime(label_fmt)
                if i % tick_every == 0 or i == len(date_range) - 1
                else ""
            )
            data.append(round(running, 2))

    # Текущий нереализованный P&L (для отображения в тултипе / в заголовке)
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    unrealized = round(
        sum(
            (float(b.last_price or b.buy_price) - float(b.buy_price)) * b.amount
            for b in active
        ),
        2,
    )

    return jsonify(
        {"labels": labels, "data": data, "unrealized": unrealized, "period": period}
    )


# ── Уведомления: ближайшие купоны ────────────────────────────────────────────


@portfolio_bp.route("/api/notifications/upcoming", methods=["GET"])
@login_required
def upcoming_notifications():
    """Возвращает купонные выплаты в ближайшие N дней (для колокольчика).

    ?days=7  — горизонт (по умолчанию 7)
    """
    try:
        days = max(1, min(int(request.args.get("days", 7)), 90))
    except (ValueError, TypeError):
        days = 7

    today = date.today()
    horizon = today + timedelta(days=days)

    active_bonds = BondPortfolio.query.filter_by(
        user_id=current_user.id, is_sold=False
    ).all()

    events: list[dict] = []
    for bond in active_bonds:
        try:
            coupons = get_coupon_calendar_cached(bond.isin) or []
        except Exception:
            continue
        for c in coupons:
            coupon_date_str = c.get("coupondate") or c.get("date") or ""
            if not coupon_date_str:
                continue
            try:
                coupon_date = date.fromisoformat(coupon_date_str[:10])
            except ValueError:
                continue
            if today <= coupon_date <= horizon:
                events.append(
                    {
                        "isin": bond.isin,
                        "name": bond.name or bond.isin,
                        "coupon_date": coupon_date_str[:10],
                        "coupon_value": c.get("value") or c.get("couponvalue"),
                        "amount": bond.amount,
                        "days_left": (coupon_date - today).days,
                    }
                )

    events.sort(key=lambda x: x["coupon_date"])
    return jsonify({"count": len(events), "events": events})


# ── Статистика (P&L по месяцам) ───────────────────────────────────────────────


@portfolio_bp.route("/api/portfolio_stats", methods=["GET"])
@login_required
def portfolio_stats():
    cache_key = f"portfolio_stats:{current_user.id}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)

    closed = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True).all()
    monthly = calc_monthly_profit(closed)
    sorted_months = sorted(monthly.keys())
    result = {
        "labels": sorted_months,
        "datasets": [
            {
                "label": "Чистая зафиксированная прибыль (₽)",
                "data": [monthly[m] for m in sorted_months],
                "backgroundColor": "rgba(40, 167, 69, 0.2)",
                "borderColor": "rgba(40, 167, 69, 1)",
                "borderWidth": 2,
                "fill": True,
            }
        ],
    }
    cache.set(cache_key, result, timeout=STATS_TTL)
    return jsonify(result)


@portfolio_bp.route("/api/portfolio/import", methods=["POST"])
@login_required
def import_portfolio():
    """Импортирует сделки из брокерского отчёта (CSV/XLSX) или JSON-массива.

    Поддерживаемые форматы:
    - Т-Инвестиции (broker=tinkoff): стандартный XLSX-отчёт по сделкам
    - Авто (broker=auto): универсальный парсер, ищет заголовки сам
    - JSON-массив: прямая передача данных через API

    Принципиально НЕ вызывает MOEX API во время импорта — это исключает
    таймаут воркера (gunicorn default 30 s) при большом числе ISINs.
    Актуальные цены/данные MOEX подтянутся при следующей загрузке портфеля.
    """
    # ── читаем брокера из формы ────────────────────────────────────────────────
    broker = (request.form.get("broker") or "auto").strip().lower()
    # Т-Инвестиции: фильтруем РЕПО-строки дополнительно по типу сделки
    filter_repo = broker in ("tinkoff", "tbank", "auto")

    # ── чтение файла ──────────────────────────────────────────────────────────
    deals = []
    skipped_repo = 0

    if "file" in request.files:
        f = request.files["file"]
        filename = (f.filename or "").lower()

        # ── XLSX ──────────────────────────────────────────────────────────────
        if filename.endswith((".xlsx", ".xls")):
            try:
                # Т-Инвестиции генерирует файлы с <dimension ref="A1"/> вместо
                # реального диапазона, поэтому read_only=True (SAX-streaming)
                # останавливается после первой строки. Используем обычный режим.
                file_bytes = io.BytesIO(f.read())
                wb = openpyxl.load_workbook(file_bytes, data_only=True)

                # Ищем нужный лист: первый, или с ключевым словом «сделки»/«trades»
                sheet = wb.active
                for sh in wb.worksheets:
                    sn = sh.title.lower()
                    if any(k in sn for k in ("сделк", "trade", "операц")):
                        sheet = sh
                        break

                # Один проход: собираем все строки как кортежи raw-значений.
                # Tuple значительно легче Cell-объектов openpyxl (нет стилей, формул).
                all_rows = list(sheet.iter_rows(values_only=True))
                wb.close()

                # ── Авто-определение брокера ───────────────────────────────
                effective_broker = broker
                if broker == "auto":
                    effective_broker = _detect_broker(all_rows)
                    # Уточняем фильтр РЕПО под обнаруженного брокера
                    if effective_broker == "tinkoff":
                        filter_repo = True

                # ── ВТБ: специализированный парсер ─────────────────────────
                if effective_broker == "vtb":
                    deals.extend(_parse_vtb_xlsx(all_rows))
                    if not deals:
                        return jsonify(
                            {
                                "status": "error",
                                "message": (
                                    "ВТБ: сделки с облигациями не найдены. "
                                    "Убедитесь, что файл содержит раздел "
                                    "«Заключённые/Завершённые сделки» с покупками "
                                    "или продажами облигаций."
                                ),
                            }
                        ), 400
                else:

                    header_row_idx, hdrs = _find_header_row_from_list(all_rows)
                    if not hdrs:
                        return jsonify(
                            {
                                "status": "error",
                                "message": (
                                    "Не найдена строка заголовков. "
                                    "Убедитесь, что отчёт содержит столбцы: "
                                    "«Код актива» (ISIN), «Количество», «Цена за единицу»."
                                ),
                            }
                        ), 400

                    isin_col = _find_col(hdrs, _ISIN)
                    amt_col = _find_col(hdrs, _AMT)
                    price_col = _find_col(hdrs, _PRICE)
                    date_col = _find_col(hdrs, _DATE)
                    type_col = _find_col(hdrs, _TYPE)
                    name_col = _find_col(hdrs, _NAME)
                    comm_col = _find_col(hdrs, _COMM)
                    curr_col = _find_col(hdrs, _CURR)
                    status_col = _find_col(hdrs, _STATUS)
                    price_curr_col = _find_col(hdrs, _PRICE_CURR)
                    deal_no_col = _find_col(hdrs, _DEAL_NO)
                    is_tinkoff = broker in ("tinkoff", "tbank")
                    seen_deals: set = set()

                    if not isin_col or not amt_col or not price_col:
                        missing = []
                        if not isin_col:
                            missing.append("ISIN / Код актива")
                        if not amt_col:
                            missing.append("Количество")
                        if not price_col:
                            missing.append("Цена за единицу")
                        return jsonify(
                            {
                                "status": "error",
                                "message": f"Не найдены обязательные столбцы: {', '.join(missing)}.",
                            }
                        ), 400

                    # Вспомогательная: безопасно достать значение по 1-based col_idx из кортежа
                    def _gc(rv, col):
                        if not col or col > len(rv):
                            return None
                        return rv[col - 1]

                    for row_values in all_rows[header_row_idx + 1 :]:
                        isin_v = _gc(row_values, isin_col)
                        if isin_v is None:
                            continue
                        isin_s = str(isin_v).strip().upper()
                        if not _is_valid_isin(isin_s):
                            continue  # пропускаем тикеры акций, РЕПО-тикеры и т.д.

                        # Т-Инвестиции: фильтр акций по валюте цены (% = облигация, RUB = акция)
                        _pc_v = (
                            str(_gc(row_values, price_curr_col) or "").strip()
                            if price_curr_col
                            else ""
                        )
                        if is_tinkoff and _pc_v.upper() == "RUB":
                            continue  # акция или инструмент с ценой в рублях — пропускаем

                        # Т-Инвестиции: дедупликация OTC-сделок (RFP + DFP = одна сделка)
                        if is_tinkoff:
                            _dc = deal_no_col or 1
                            _dn = _gc(row_values, _dc)
                            _dk = str(_dn).strip() if _dn is not None else ""
                            if _dk and _dk in seen_deals:
                                continue
                            if _dk:
                                seen_deals.add(_dk)

                        # Пропускаем РЕПО-сделки
                        type_v = _gc(row_values, type_col)
                        if filter_repo and _is_repo(type_v):
                            skipped_repo += 1
                            continue

                        # Пропускаем отменённые
                        if status_col and _is_cancelled(_gc(row_values, status_col)):
                            continue

                        raw_curr = str(_gc(row_values, curr_col) or "").strip().upper()
                        currency = (
                            raw_curr if raw_curr.isalpha() and len(raw_curr) == 3 else "RUB"
                        )

                        # Т-Инвестиции: цена в % от номинала (номинал = 1000 ₽) → рублей
                        _price_v = _gc(row_values, price_col)
                        if is_tinkoff and _pc_v == "%" and _price_v is not None:
                            try:
                                _price_v = float(_price_v) * 10
                            except (TypeError, ValueError):
                                pass

                        name_v = _gc(row_values, name_col)
                        deals.append(
                            {
                                "isin": isin_s,
                                "amount": _gc(row_values, amt_col),
                                "price": _price_v,
                                "date": _gc(row_values, date_col),
                                "tx_type": _tx_type(type_v),
                                "name": str(name_v).strip() if name_v else None,
                                "commission": _gc(row_values, comm_col),
                                "currency": currency,
                                "notes": "",
                            }
                        )

                    # Т-Инвестиции: купонные выплаты из Раздела 2 отчёта
                    if is_tinkoff:
                        _re_isin = re.compile(r"ISIN:\s*([A-Z0-9]{12})", re.IGNORECASE)
                        _re_qty = re.compile(r"[Кк]оличество[^:]*:\s*(\d+)")
                        _re_punit = re.compile(
                            r"(?:купоны за 1 бумагу|за 1 ценную бумагу)[^:]*:\s*([\d,.]+)"
                        )
                        _re_date = re.compile(
                            r"Дата операции:\s*(\d{2})-([A-Za-z]{3})-(\d{2,4})"
                        )
                        _MON = {
                            m: i
                            for i, m in enumerate(
                                [
                                    "JAN",
                                    "FEB",
                                    "MAR",
                                    "APR",
                                    "MAY",
                                    "JUN",
                                    "JUL",
                                    "AUG",
                                    "SEP",
                                    "OCT",
                                    "NOV",
                                    "DEC",
                                ],
                                1,
                            )
                        }
                        # Сканируем все строки (all_rows уже в памяти — повторный проход бесплатен)
                        for row_values in all_rows:
                            desc = next(
                                (
                                    str(v)
                                    for v in row_values
                                    if v
                                    and "isin" in str(v).lower()
                                    and "купон" in str(v).lower()
                                ),
                                None,
                            )
                            if not desc:
                                continue
                            m_isin = _re_isin.search(desc)
                            m_qty = _re_qty.search(desc)
                            m_punit = _re_punit.search(desc)
                            if not (m_isin and m_qty and m_punit):
                                continue
                            c_isin = m_isin.group(1).upper()
                            c_qty = int(m_qty.group(1))
                            c_punit = float(m_punit.group(1).replace(",", "."))
                            c_date = date.today()
                            m_date = _re_date.search(desc)
                            if m_date:
                                try:
                                    day = int(m_date.group(1))
                                    mon = _MON.get(m_date.group(2).upper(), 1)
                                    yr = int(m_date.group(3))
                                    c_date = date(2000 + yr if yr < 100 else yr, mon, day)
                                except (ValueError, KeyError):
                                    pass
                            deals.append(
                                {
                                    "isin": c_isin,
                                    "amount": c_qty,
                                    "price": c_punit,
                                    "date": c_date,
                                    "tx_type": "coupon",
                                    "name": None,
                                    "commission": None,
                                    "currency": "RUB",
                                    "notes": "Купонный доход (Т-Инвестиции)",
                                }
                            )

            except Exception as exc:
                logger.error("XLSX import parse error: %s", exc, exc_info=True)
                return jsonify(
                    {"status": "error", "message": f"Ошибка обработки XLSX: {exc}"}
                ), 400

        # ── CSV ───────────────────────────────────────────────────────────────
        elif filename.endswith(".csv"):
            try:
                raw = f.read()
                text = None
                for enc in ("utf-8-sig", "cp1251", "utf-8"):
                    try:
                        text = raw.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                if text is None:
                    text = raw.decode("utf-8", errors="replace")

                lines = text.splitlines()
                first_line = lines[0] if lines else ""
                delim = (
                    "\t" if "\t" in first_line else (";" if ";" in first_line else ",")
                )

                # Ищем строку с заголовками
                header_idx = 0
                for i, line in enumerate(lines):
                    ln = line.lower()
                    if any(k in ln for k in ["isin", "код актива", "код бумаги"]):
                        header_idx = i
                        break

                import csv as _csv

                reader = _csv.DictReader(
                    [lines[header_idx]] + lines[header_idx + 1 :],
                    delimiter=delim,
                )

                def _csv_get(rn, candidates):
                    for name in candidates:
                        v = rn.get(_norm_hdr(name))
                        if v is not None and str(v).strip():
                            return v
                    return None

                for raw_row in reader:
                    rn = {_norm_hdr(k): v for k, v in raw_row.items()}
                    isin_v = _csv_get(rn, _ISIN)
                    if not isin_v:
                        continue
                    isin_s = str(isin_v).strip().upper()
                    if not _is_valid_isin(isin_s):
                        continue
                    type_v = _csv_get(rn, _TYPE)
                    if filter_repo and _is_repo(type_v):
                        skipped_repo += 1
                        continue
                    if _is_cancelled(_csv_get(rn, _STATUS)):
                        continue

                    raw_curr = str(_csv_get(rn, _CURR) or "").strip().upper()
                    currency = (
                        raw_curr if raw_curr.isalpha() and len(raw_curr) == 3 else "RUB"
                    )

                    deals.append(
                        {
                            "isin": isin_s,
                            "amount": _csv_get(rn, _AMT),
                            "price": _csv_get(rn, _PRICE),
                            "date": _csv_get(rn, _DATE),
                            "tx_type": _tx_type(type_v),
                            "name": _csv_get(rn, _NAME),
                            "commission": _csv_get(rn, _COMM),
                            "currency": currency,
                            "notes": "",
                        }
                    )
            except Exception as exc:
                logger.error("CSV import parse error: %s", exc, exc_info=True)
                return jsonify(
                    {"status": "error", "message": f"Ошибка обработки CSV: {exc}"}
                ), 400

        else:
            return jsonify(
                {"status": "error", "message": "Поддерживаются только .csv и .xlsx"}
            ), 400

    else:
        deals = (request.get_json() or {}).get("deals", [])

    if not deals:
        hint = ""
        if skipped_repo:
            hint = f" (отфильтровано РЕПО-сделок: {skipped_repo})"
        return jsonify(
            {
                "status": "error",
                "message": (
                    f"Сделки с облигациями не найдены{hint}. "
                    "Убедитесь, что файл содержит покупки/продажи облигаций."
                ),
            }
        ), 400

    # ── обработка сделок — БЕЗ вызовов MOEX API (предотвращает таймаут) ───────
    # Цены/secid обновятся автоматически при следующей загрузке портфеля.
    imported_count = 0
    coupon_count = 0
    errors: list = []

    for deal in deals:
        isin = str(deal.get("isin", "")).strip().upper()
        tx_type = deal.get("tx_type", "buy")
        notes = deal.get("notes") or ""

        if not isin:
            continue

        raw_amt = _parse_num(deal.get("amount"))
        if raw_amt is None or raw_amt <= 0:
            errors.append(
                f"Пропущено {isin}: некорректное количество ({deal.get('amount')!r})"
            )
            continue
        amount = int(raw_amt)

        price = _parse_num(deal.get("price"))
        if price is None or price <= 0:
            errors.append(
                f"Пропущено {isin}: некорректная цена ({deal.get('price')!r})"
            )
            continue

        trade_date = _parse_any_date(deal.get("date"))
        commission = _parse_num(deal.get("commission"))
        currency = deal.get("currency") or "RUB"
        bond_title = deal.get("name") or isin

        if tx_type == "buy":
            db.session.add(
                BondPortfolio(
                    user_id=current_user.id,
                    isin=isin,
                    secid=isin,  # обновится при загрузке портфеля
                    name=bond_title,
                    amount=amount,
                    buy_price=price,
                    last_price=price,  # обновится при загрузке портфеля
                    purchase_date=trade_date,
                    is_sold=False,
                    currency=currency,
                    broker_commission=commission,
                    notes=notes or None,
                )
            )
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="buy",
                    amount=amount,
                    price=price,
                    commission=commission,
                    currency=currency,
                    tx_date=trade_date,
                )
            )
            imported_count += 1

        elif tx_type == "coupon":
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="coupon",
                    amount=amount,
                    price=price,
                    commission=None,
                    currency=currency,
                    tx_date=trade_date,
                )
            )
            coupon_count += 1
            imported_count += 1

        else:  # sell
            active = BondPortfolio.query.filter_by(
                user_id=current_user.id, isin=isin, is_sold=False
            ).first()
            if active:
                active.is_sold = True
                active.sell_price = price
                active.sell_date = trade_date
                if commission:
                    active.broker_commission = commission
            else:
                db.session.add(
                    BondPortfolio(
                        user_id=current_user.id,
                        isin=isin,
                        secid=isin,
                        name=bond_title,
                        amount=amount,
                        buy_price=price,
                        last_price=price,
                        purchase_date=trade_date,
                        is_sold=True,
                        sell_price=price,
                        sell_date=trade_date,
                        currency=currency,
                        broker_commission=commission,
                    )
                )
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="sell",
                    amount=amount,
                    price=price,
                    commission=commission,
                    currency=currency,
                    tx_date=trade_date,
                )
            )
            imported_count += 1

    db.session.commit()
    _bust_user_cache(current_user.id)

    msg = f"Импортировано {imported_count} записей."
    if coupon_count:
        msg += f" Купонных выплат: {coupon_count}."
    if skipped_repo:
        msg += f" РЕПО-сделок пропущено: {skipped_repo}."
    if errors:
        msg += f" Ошибок: {len(errors)}."
    return jsonify(
        {
            "status": "success",
            "message": msg,
            "imported_count": imported_count,
            "coupon_count": coupon_count,
            "errors": errors,
        }
    ), 200
