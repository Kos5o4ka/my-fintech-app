import csv
import io
import json
import logging
import openpyxl
import re
from datetime import datetime, date
from typing import Optional
from openpyxl.styles import Font, PatternFill, Alignment
from io import StringIO

from flask import Blueprint, request, jsonify, render_template, make_response
from flask_login import login_required, current_user

from extensions import db, cache
from models import BondPortfolio, Transaction

logger = logging.getLogger(__name__)
imports_bp = Blueprint("imports", __name__)

# ── Список псевдонимов столбцов для парсинга брокерских отчётов ───────────────
_ISIN = ["isin", "isin код", "код актива", "код бумаги", "код инструмента", "code"]
_AMT = ["количество", "кол-во", "кол.", "amount", "qty", "объем", "объём"]
_PRICE = ["цена за единицу", "цена сделки", "цена", "price", "курс"]
_DATE = [
    "дата заключения",
    "дата сделки",
    "дата",
    "date",
    "дата операции",
    "дата торгов",
]
_TYPE = ["вид сделки", "тип сделки", "операция", "тип операции", "направление", "type"]
_NAME = [
    "наименование актива",
    "наименование инструмента",
    "наименование",
    "название",
    "инструмент",
    "name",
]
_COMM = [
    "комиссия брокера",
    "сумма комиссии брокера",
    "комиссия",
    "commission",
    "broker_commission",
]
_CURR = ["валюта расчетов", "валюта расчётов", "валюта", "currency"]
_STATUS = ["признак исполнения", "статус", "status"]
_PRICE_CURR = ["валюта цены", "единица цены", "валюта цены сделки"]
_DEAL_NO = ["номер сделки", "№ сделки", "n сделки", "номер"]
_ANCHORS = _ISIN + _AMT + _PRICE


def _norm_hdr(v) -> str:
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").replace("\r", " ").split()).lower()


def _parse_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = (
        str(v)
        .strip()
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _parse_any_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return date.today()
    s = str(v).strip().split()[0].split("/")[0].strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.today()


def _is_valid_isin(s: str) -> bool:
    return len(s) == 12 and s.isalnum() and s[:2].isalpha()


def _tx_type(val) -> str:
    if not val:
        return "buy"
    v = str(val).strip().lower()
    if v in ("продажа", "sell", "s", "-", "реализация", "погашение"):
        return "sell"
    return "buy"


def _is_repo(val) -> bool:
    if not val:
        return False
    return "репо" in str(val).lower() or "repo" in str(val).lower()


def _is_cancelled(val) -> bool:
    if not val:
        return False
    v = str(val).strip().lower()
    return v in ("отменена", "отменено", "cancelled", "canceled", "rejected")


def _find_header_row_xlsx(sheet, max_scan: int = 50):
    anchors = {_norm_hdr(a) for a in _ANCHORS}
    max_r = sheet.max_row or max_scan
    for ri in range(1, min(max_scan + 1, max_r + 1)):
        hdrs = {}
        for cell in sheet[ri]:
            n = _norm_hdr(cell.value)
            if n:
                hdrs[n] = cell.column
        if sum(1 for a in anchors if a in hdrs) >= 2:
            return ri, hdrs
    return None, {}


def _find_header_row_from_list(rows: list, max_scan: int = 50):
    anchors = {_norm_hdr(a) for a in _ANCHORS}
    for ri, row_values in enumerate(rows[:max_scan]):
        hdrs = {}
        for col_idx, cell_val in enumerate(row_values, start=1):
            n = _norm_hdr(cell_val)
            if n:
                hdrs[n] = col_idx
        if sum(1 for a in anchors if a in hdrs) >= 2:
            return ri, hdrs
    return None, {}


def _find_col(hdrs: dict, candidates: list):
    for name in candidates:
        n = _norm_hdr(name)
        if n in hdrs:
            return hdrs[n]
    for name in candidates:
        n = _norm_hdr(name)
        for hn, hc in hdrs.items():
            if hn.startswith(n) or n.startswith(hn):
                return hc
    return None


def _parse_vtb_xlsx(all_rows: list) -> list:
    seen_deals = set()
    deals = []

    for row in all_rows:
        if not row or len(row) < 18:
            continue

        type_cell = row[5]
        if not type_cell:
            continue
        type_s = str(type_cell).strip().lower()
        if type_s not in ("покупка", "продажа"):
            continue

        name_cell = row[1]
        if not name_cell:
            continue

        parts = [p.strip() for p in str(name_cell).split(",")]
        if len(parts) < 3:
            continue

        isin = parts[-1].upper()
        if not _is_valid_isin(isin):
            continue

        reg_code = parts[-2].upper()
        if not (reg_code.startswith("4B") or "RMFS" in reg_code):
            continue

        deal_no = (
            str(row[25]).strip()
            if len(row) > 25 and row[25] is not None
            else ""
        )
        if deal_no and deal_no in seen_deals:
            continue
        if deal_no:
            seen_deals.add(deal_no)

        price_pct = row[9]
        qty_v = row[7]
        date_v = row[2]
        comm1 = row[15] if len(row) > 15 else None
        comm2 = row[17] if len(row) > 17 else None
        curr_v = row[11] if len(row) > 11 else None

        try:
            price_pct_val = float(price_pct)
        except (TypeError, ValueError):
            continue
        if price_pct_val <= 0:
            continue

        commission = 0.0
        for c in (comm1, comm2):
            try:
                commission += float(c)
            except (TypeError, ValueError):
                pass

        raw_curr = str(curr_v or "").strip().upper()
        if raw_curr in ("RUR", "RUB"):
            currency = "RUB"
        elif raw_curr.isalpha() and len(raw_curr) == 3:
            currency = raw_curr
        else:
            currency = "RUB"

        deals.append(
            {
                "isin": isin,
                "amount": qty_v,
                "price": price_pct_val,  # сохраняем в процентах, починится через self-healing
                "date": date_v,
                "tx_type": "sell" if type_s == "продажа" else "buy",
                "name": ", ".join(parts[:-2]),
                "commission": commission if commission > 0 else None,
                "currency": currency,
                "notes": None,
                "deal_no": deal_no,
            }
        )

    return deals


def _detect_broker(all_rows: list) -> str:
    vtb_votes = 0
    for row in all_rows[:100]:
        if not row or len(row) < 6:
            continue
        if str(row[5] or "").strip().lower() in ("покупка", "продажа"):
            if str(row[1] or "").count(",") >= 2:
                vtb_votes += 1
                if vtb_votes >= 2:
                    return "vtb"

    tinkoff_markers = ("т-инвестиции", "tinkoff", "tbank", "т инвестиции")
    for row in all_rows[:40]:
        if not row:
            continue
        for cell in row:
            if not cell:
                continue
            s = str(cell).strip().lower()
            if any(m in s for m in tinkoff_markers):
                return "tinkoff"

    return "generic"


def _bust_user_cache(user_id: int) -> None:
    cache.delete(f"portfolio_stats:{user_id}")
    cache.delete(f"portfolio_calendar:{user_id}")
    cache.delete(f"portfolio_income:{user_id}")
    cache.delete(f"portfolio_sharpe:{user_id}")
    cache.delete("moex_currency_rates")


@imports_bp.route("/import")
@login_required
def import_page():
    """Страница импорта отчетов."""
    return render_template("import.html", active_page="import")


@imports_bp.route("/api/portfolio/import", methods=["POST"])
@login_required
def import_portfolio():
    """Импортирует сделки из отчета брокера с интеллектуальной дедупликацией."""
    broker = (request.form.get("broker") or "auto").strip().lower()
    filter_repo = broker in ("tinkoff", "tbank", "auto")

    deals = []
    skipped_repo = 0

    if "file" in request.files:
        f = request.files["file"]
        filename = (f.filename or "").lower()

        # ── XLSX ──────────────────────────────────────────────────────────────
        if filename.endswith((".xlsx", ".xls")):
            try:
                file_bytes = io.BytesIO(f.read())
                wb = openpyxl.load_workbook(file_bytes, data_only=True)

                sheet = wb.active
                for sh in wb.worksheets:
                    sn = sh.title.lower()
                    if any(k in sn for k in ("сделк", "trade", "операц")):
                        sheet = sh
                        break

                all_rows = list(sheet.iter_rows(values_only=True))
                wb.close()

                effective_broker = broker
                if broker == "auto":
                    effective_broker = _detect_broker(all_rows)
                    if effective_broker == "tinkoff":
                        filter_repo = True

                if effective_broker == "vtb":
                    deals.extend(_parse_vtb_xlsx(all_rows))
                    if not deals:
                        return jsonify(
                            {
                                "status": "error",
                                "message": (
                                    "ВТБ: сделки с облигациями не найдены. "
                                    "Убедитесь, что файл содержит раздел "
                                    "«Заключённые/Завершённые сделки» с покупками "
                                    "или продажами облигаций."
                                ),
                            }
                        ), 400
                else:
                    header_row_idx, hdrs = _find_header_row_from_list(all_rows)
                    if not hdrs:
                        return jsonify(
                            {
                                "status": "error",
                                "message": (
                                    "Не найдена строка заголовков. "
                                    "Убедитесь, что отчёт содержит столбцы: "
                                    "«Код актива» (ISIN), «Количество», «Цена за единицу»."
                                ),
                            }
                        ), 400

                    isin_col = _find_col(hdrs, _ISIN)
                    amt_col = _find_col(hdrs, _AMT)
                    price_col = _find_col(hdrs, _PRICE)
                    date_col = _find_col(hdrs, _DATE)
                    type_col = _find_col(hdrs, _TYPE)
                    name_col = _find_col(hdrs, _NAME)
                    comm_col = _find_col(hdrs, _COMM)
                    curr_col = _find_col(hdrs, _CURR)
                    status_col = _find_col(hdrs, _STATUS)
                    price_curr_col = _find_col(hdrs, _PRICE_CURR)
                    deal_no_col = _find_col(hdrs, _DEAL_NO)
                    is_tinkoff = effective_broker in ("tinkoff", "tbank")
                    seen_deals = set()

                    if not isin_col or not amt_col or not price_col:
                        missing = []
                        if not isin_col:
                            missing.append("ISIN / Код актива")
                        if not amt_col:
                            missing.append("Количество")
                        if not price_col:
                            missing.append("Цена за единицу")
                        return jsonify(
                            {
                                "status": "error",
                                "message": f"Не найдены обязательные столбцы: {', '.join(missing)}.",
                            }
                        ), 400

                    def _gc(rv, col):
                        if not col or col > len(rv):
                            return None
                        return rv[col - 1]

                    for row_values in all_rows[header_row_idx + 1 :]:
                        isin_v = _gc(row_values, isin_col)
                        if isin_v is None:
                            continue
                        isin_s = str(isin_v).strip().upper()
                        if not _is_valid_isin(isin_s):
                            continue

                        _pc_v = (
                            str(_gc(row_values, price_curr_col) or "").strip()
                            if price_curr_col
                            else ""
                        )
                        if is_tinkoff and _pc_v.upper() == "RUB":
                            continue

                        deal_no = ""
                        if is_tinkoff:
                            _dc = deal_no_col or 1
                            _dn = _gc(row_values, _dc)
                            _dk = str(_dn).strip() if _dn is not None else ""
                            if _dk and _dk in seen_deals:
                                continue
                            if _dk:
                                seen_deals.add(_dk)
                                deal_no = _dk
                        elif deal_no_col:
                            _dn = _gc(row_values, deal_no_col)
                            deal_no = str(_dn).strip() if _dn is not None else ""

                        type_v = _gc(row_values, type_col)
                        if filter_repo and _is_repo(type_v):
                            skipped_repo += 1
                            continue

                        if status_col and _is_cancelled(_gc(row_values, status_col)):
                            continue

                        raw_curr = str(_gc(row_values, curr_col) or "").strip().upper()
                        currency = (
                            raw_curr if raw_curr.isalpha() and len(raw_curr) == 3 else "RUB"
                        )

                        _price_v = _gc(row_values, price_col)
                        price_val = _parse_num(_price_v)

                        name_v = _gc(row_values, name_col)
                        deals.append(
                            {
                                "isin": isin_s,
                                "amount": _gc(row_values, amt_col),
                                "price": price_val,
                                "date": _gc(row_values, date_col),
                                "tx_type": _tx_type(type_v),
                                "name": str(name_v).strip() if name_v else None,
                                "commission": _gc(row_values, comm_col),
                                "currency": currency,
                                "notes": "",
                                "deal_no": deal_no,
                            }
                        )

                    # Т-Инвестиции: купонные выплаты из Раздела 2 отчёта
                    if is_tinkoff:
                        _re_isin = re.compile(r"ISIN:\s*([A-Z0-9]{12})", re.IGNORECASE)
                        _re_qty = re.compile(r"[Кк]оличество[^:]*:\s*(\d+)")
                        _re_punit = re.compile(
                            r"(?:купоны за 1 бумагу|за 1 ценную бумагу)[^:]*:\s*([\d,.]+)"
                        )
                        _re_date = re.compile(
                            r"Дата операции:\s*(\d{2})-([A-Za-z]{3})-(\d{2,4})"
                        )
                        _MON = {
                            m: i
                            for i, m in enumerate(
                                [
                                    "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                                    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
                                ],
                                1,
                            )
                        }
                        for row_values in all_rows:
                            desc = next(
                                (
                                    str(v)
                                    for v in row_values
                                    if v
                                    and "isin" in str(v).lower()
                                    and "купон" in str(v).lower()
                                ),
                                None,
                            )
                            if not desc:
                                continue
                            m_isin = _re_isin.search(desc)
                            m_qty = _re_qty.search(desc)
                            m_punit = _re_punit.search(desc)
                            if not (m_isin and m_qty and m_punit):
                                continue
                            c_isin = m_isin.group(1).upper()
                            c_qty = int(m_qty.group(1))
                            c_punit = float(m_punit.group(1).replace(",", "."))
                            c_date = date.today()
                            m_date = _re_date.search(desc)
                            if m_date:
                                try:
                                    day = int(m_date.group(1))
                                    mon = _MON.get(m_date.group(2).upper(), 1)
                                    yr = int(m_date.group(3))
                                    c_date = date(2000 + yr if yr < 100 else yr, mon, day)
                                    reward_deal_no = f"coupon-{c_isin}-{c_date.strftime('%Y%m%d')}-{c_qty}"
                                except (ValueError, KeyError):
                                    reward_deal_no = ""
                            else:
                                reward_deal_no = ""

                            deals.append(
                                {
                                    "isin": c_isin,
                                    "amount": c_qty,
                                    "price": c_punit,
                                    "date": c_date,
                                    "tx_type": "coupon",
                                    "name": None,
                                    "commission": None,
                                    "currency": "RUB",
                                    "notes": "Купонный доход (Т-Инвестиции)",
                                    "deal_no": reward_deal_no,
                                }
                            )

            except Exception as exc:
                logger.error("XLSX import parse error: %s", exc, exc_info=True)
                return jsonify(
                    {"status": "error", "message": f"Ошибка обработки XLSX: {exc}"}
                ), 400

        # ── CSV ───────────────────────────────────────────────────────────────
        elif filename.endswith(".csv"):
            try:
                raw = f.read()
                text = None
                for enc in ("utf-8-sig", "cp1251", "utf-8"):
                    try:
                        text = raw.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                if text is None:
                    text = raw.decode("utf-8", errors="replace")

                lines = text.splitlines()
                first_line = lines[0] if lines else ""
                delim = (
                    "\t" if "\t" in first_line else (";" if ";" in first_line else ",")
                )

                header_idx = 0
                for i, line in enumerate(lines):
                    ln = line.lower()
                    if any(k in ln for k in ["isin", "код актива", "код бумаги"]):
                        header_idx = i
                        break

                reader = csv.DictReader(
                    [lines[header_idx]] + lines[header_idx + 1 :],
                    delimiter=delim,
                )

                def _csv_get(rn, candidates):
                    for name in candidates:
                        v = rn.get(_norm_hdr(name))
                        if v is not None and str(v).strip():
                            return v
                    return None

                for raw_row in reader:
                    rn = {_norm_hdr(k): v for k, v in raw_row.items()}
                    isin_v = _csv_get(rn, _ISIN)
                    if not isin_v:
                        continue
                    isin_s = str(isin_v).strip().upper()
                    if not _is_valid_isin(isin_s):
                        continue
                    type_v = _csv_get(rn, _TYPE)
                    if filter_repo and _is_repo(type_v):
                        skipped_repo += 1
                        continue
                    if _is_cancelled(_csv_get(rn, _STATUS)):
                        continue

                    raw_curr = str(_csv_get(rn, _CURR) or "").strip().upper()
                    currency = (
                        raw_curr if raw_curr.isalpha() and len(raw_curr) == 3 else "RUB"
                    )

                    deal_no = _csv_get(rn, _DEAL_NO) or ""

                    deals.append(
                        {
                            "isin": isin_s,
                            "amount": _csv_get(rn, _AMT),
                            "price": _csv_get(rn, _PRICE),
                            "date": _csv_get(rn, _DATE),
                            "tx_type": _tx_type(type_v),
                            "name": _csv_get(rn, _NAME),
                            "commission": _csv_get(rn, _COMM),
                            "currency": currency,
                            "notes": "",
                            "deal_no": deal_no,
                        }
                    )
            except Exception as exc:
                logger.error("CSV import parse error: %s", exc, exc_info=True)
                return jsonify(
                    {"status": "error", "message": f"Ошибка обработки CSV: {exc}"}
                ), 400

        else:
            return jsonify(
                {"status": "error", "message": "Поддерживаются только .csv и .xlsx"}
            ), 400

    else:
        deals = (request.get_json() or {}).get("deals", [])

    if not deals:
        hint = ""
        if skipped_repo:
            hint = f" (отфильтровано РЕПО-сделок: {skipped_repo})"
        return jsonify(
            {
                "status": "error",
                "message": (
                    f"Сделки с облигациями не найдены{hint}. "
                    "Убедитесь, что файл содержит покупки/продажи облигаций."
                ),
            }
        ), 400

    imported_count = 0
    coupon_count = 0
    errors = []

    for deal in deals:
        isin = str(deal.get("isin", "")).strip().upper()
        tx_type = deal.get("tx_type", "buy")
        notes = deal.get("notes") or ""
        deal_no = str(deal.get("deal_no") or "").strip()

        if not isin:
            continue

        raw_amt = _parse_num(deal.get("amount"))
        if raw_amt is None or raw_amt <= 0:
            errors.append(
                f"Пропущено {isin}: некорректное количество ({deal.get('amount')!r})"
            )
            continue
        amount = int(raw_amt)

        price = _parse_num(deal.get("price"))
        if price is None or price <= 0:
            errors.append(
                f"Пропущено {isin}: некорректная цена ({deal.get('price')!r})"
            )
            continue

        trade_date = _parse_any_date(deal.get("date"))
        commission = _parse_num(deal.get("commission"))
        currency = deal.get("currency") or "RUB"
        bond_title = deal.get("name") or isin

        # ── ИНТЕЛЛЕКТУАЛЬНАЯ ДЕДУПЛИКАЦИЯ ──
        if deal_no:
            exists = Transaction.query.filter_by(
                user_id=current_user.id, deal_no=deal_no
            ).first()
            if exists:
                continue  # Пропускаем дубликат!
        else:
            exists = Transaction.query.filter_by(
                user_id=current_user.id,
                isin=isin,
                tx_type=tx_type,
                amount=amount,
                price=price,
                tx_date=trade_date,
                deal_no=None,
            ).first()
            if exists:
                continue  # Пропускаем дубликат!

        if tx_type == "buy":
            db.session.add(
                BondPortfolio(
                    user_id=current_user.id,
                    isin=isin,
                    secid=isin,
                    name=bond_title,
                    amount=amount,
                    buy_price=price,
                    last_price=price,
                    purchase_date=trade_date,
                    is_sold=False,
                    currency=currency,
                    broker_commission=commission,
                    notes=notes or None,
                    buy_deal_no=deal_no or None,
                )
            )
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="buy",
                    amount=amount,
                    price=price,
                    commission=commission,
                    currency=currency,
                    tx_date=trade_date,
                    deal_no=deal_no or None,
                )
            )
            imported_count += 1

        elif tx_type == "coupon":
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="coupon",
                    amount=amount,
                    price=price,
                    commission=None,
                    currency=currency,
                    tx_date=trade_date,
                    deal_no=deal_no or None,
                )
            )
            coupon_count += 1
            imported_count += 1

        else:  # sell
            active = BondPortfolio.query.filter_by(
                user_id=current_user.id, isin=isin, is_sold=False
            ).first()
            if active:
                active.is_sold = True
                active.sell_price = price
                active.sell_date = trade_date
                active.sell_deal_no = deal_no or None
                if commission:
                    active.broker_commission = commission
            else:
                db.session.add(
                    BondPortfolio(
                        user_id=current_user.id,
                        isin=isin,
                        secid=isin,
                        name=bond_title,
                        amount=amount,
                        buy_price=price,
                        last_price=price,
                        purchase_date=trade_date,
                        is_sold=True,
                        sell_price=price,
                        sell_date=trade_date,
                        currency=currency,
                        broker_commission=commission,
                        buy_deal_no=None,
                        sell_deal_no=deal_no or None,
                    )
                )
            db.session.add(
                Transaction(
                    user_id=current_user.id,
                    isin=isin,
                    name=bond_title,
                    tx_type="sell",
                    amount=amount,
                    price=price,
                    commission=commission,
                    currency=currency,
                    tx_date=trade_date,
                    deal_no=deal_no or None,
                )
            )
            imported_count += 1

    db.session.commit()
    _bust_user_cache(current_user.id)

    msg = f"Импортировано {imported_count} записей."
    if coupon_count:
        msg += f" Купонных выплат: {coupon_count}."
    if skipped_repo:
        msg += f" РЕПО-сделок пропущено: {skipped_repo}."
    if errors:
        msg += f" Ошибок: {len(errors)}."
    return jsonify(
        {
            "status": "success",
            "message": msg,
            "imported_count": imported_count,
            "coupon_count": coupon_count,
            "errors": errors,
        }
    ), 200


@imports_bp.route("/api/portfolio/export", methods=["GET"])
@login_required
def export_portfolio_csv():
    """Экспорт портфеля в CSV."""
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(
        [
            "Название бумаги",
            "ISIN код",
            "Количество (шт)",
            "Цена покупки (руб)",
            "Дата сделки",
        ]
    )
    for bond in active:
        cw.writerow(
            [bond.name, bond.isin, bond.amount, bond.buy_price, bond.purchase_date]
        )
    response = make_response("﻿" + si.getvalue())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.csv"
    )
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    return response


@imports_bp.route("/api/portfolio/export/xlsx", methods=["GET"])
@login_required
def export_portfolio_xlsx():
    """Экспорт портфеля и истории сделок в XLSX."""
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    sold = (
        BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True)
        .order_by(BondPortfolio.sell_date.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "Портфель"
    headers1 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Посл. цена (₽)",
        "P&L (₽)",
        "Дата покупки",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1E7E34")
        cell.alignment = center
    for bond in active:
        last_p = float(bond.last_price) if bond.last_price else float(bond.buy_price)
        pnl = round((last_p - float(bond.buy_price)) * bond.amount, 2)
        ws1.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                float(bond.buy_price),
                round(last_p, 2),
                pnl,
                bond.purchase_date.strftime("%Y-%m-%d") if bond.purchase_date else "",
            ]
        )
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    ws2 = wb.create_sheet("История сделок")
    headers2 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Цена продажи (₽)",
        "Комиссия (₽)",
        "P&L (₽)",
        "P&L %",
        "Дата продажи",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="155724")
        cell.alignment = center
    for bond in sold:
        buy_p = float(bond.buy_price)
        sell_p = float(bond.sell_price) if bond.sell_price else buy_p
        comm = float(bond.broker_commission) if bond.broker_commission else 0.0
        pnl = round((sell_p - buy_p) * bond.amount - comm, 2)
        pnl_pct = round(pnl / (buy_p * bond.amount) * 100, 2) if buy_p else 0.0
        ws2.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                round(buy_p, 2),
                round(sell_p, 2),
                round(comm, 2),
                pnl,
                pnl_pct,
                bond.sell_date.strftime("%Y-%m-%d") if bond.sell_date else "",
            ]
        )
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.xlsx"
    )
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response
