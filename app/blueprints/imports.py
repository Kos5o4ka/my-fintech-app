import csv
import io
import logging
import openpyxl
from io import StringIO

from flask import Blueprint, request, jsonify, render_template, make_response
from flask_login import login_required, current_user
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import BondPortfolio

logger = logging.getLogger(__name__)
imports_bp = Blueprint("imports", __name__)


@imports_bp.route("/import")
@login_required
def import_page():
    """Страница импорта отчетов."""
    return render_template("import.html", active_page="import")


@imports_bp.route("/api/portfolio/import", methods=["POST"])
@login_required
def import_portfolio():
    """Импортирует сделки из отчета брокера с интеллектуальной дедупликацией."""
    broker = (request.form.get("broker") or "auto").strip().lower()

    deals = []
    skipped_repo = 0
    errors = []

    if "file" in request.files:
        f = request.files["file"]
        filename = f.filename or ""

        try:
            file_content = f.read()
            from app.services.import_service import parse_broker_file

            deals, skipped_repo, err = parse_broker_file(file_content, filename, broker)
            if err:
                return jsonify({"status": "error", "message": err}), 400

        except Exception as exc:
            logger.error("File import upload failed: %s", exc)
            return jsonify(
                {"status": "error", "message": f"Ошибка чтения файла: {exc}"}
            ), 400
    else:
        deals = (request.get_json() or {}).get("deals", [])
        skipped_repo = 0

    if not deals:
        hint = ""
        if skipped_repo:
            hint = f" (отфильтровано РЕПО-сделок: {skipped_repo})"
        return jsonify(
            {
                "status": "error",
                "message": f"Сделки с облигациями не найдены{hint}. Убедитесь, что файл содержит покупки/продажи облигаций.",
            }
        ), 400

    from app.services.import_service import save_imported_deals

    try:
        imported_count, coupon_count, errors = save_imported_deals(
            deals, current_user.id
        )
    except Exception as exc:
        logger.error("save_imported_deals failed: %s", exc, exc_info=True)
        return jsonify(
            {"status": "error", "message": f"Ошибка при сохранении сделок: {exc}"}
        ), 500

    # If DB commit failed, errors contains a single critical error message
    if (
        imported_count == 0
        and coupon_count == 0
        and len(errors) == 1
        and errors[0].startswith("Ошибка сохранения")
    ):
        return jsonify({"status": "error", "message": errors[0]}), 500

    msg = f"Импортировано {imported_count} записей."
    if coupon_count:
        msg += f" Купонных выплат: {coupon_count}."
    if skipped_repo:
        msg += f" РЕПО-сделок пропущено: {skipped_repo}."
    if errors:
        msg += f" Ошибок: {len(errors)}."

    return jsonify(
        {
            "status": "success",
            "message": msg,
            "imported_count": imported_count,
            "coupon_count": coupon_count,
            "errors": errors,
        }
    ), 200


@imports_bp.route("/api/portfolio/export", methods=["GET"])
@login_required
def export_portfolio_csv():
    """Экспорт портфеля в CSV."""
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(
        [
            "Название бумаги",
            "ISIN код",
            "Количество (шт)",
            "Цена покупки (руб)",
            "Дата сделки",
        ]
    )
    for bond in active:
        cw.writerow(
            [bond.name, bond.isin, bond.amount, bond.buy_price, bond.purchase_date]
        )
    response = make_response("﻿" + si.getvalue())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.csv"
    )
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    return response


@imports_bp.route("/api/portfolio/export/xlsx", methods=["GET"])
@login_required
def export_portfolio_xlsx():
    """Экспорт портфеля и истории сделок в XLSX."""
    active = BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=False).all()
    sold = (
        BondPortfolio.query.filter_by(user_id=current_user.id, is_sold=True)
        .order_by(BondPortfolio.sell_date.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "Портфель"
    headers1 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Посл. цена (₽)",
        "P&L (₽)",
        "Дата покупки",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1E7E34")
        cell.alignment = center
    for bond in active:
        last_p = float(bond.last_price) if bond.last_price else float(bond.buy_price)
        pnl = round((last_p - float(bond.buy_price)) * bond.amount, 2)
        ws1.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                float(bond.buy_price),
                round(last_p, 2),
                pnl,
                bond.purchase_date.strftime("%Y-%m-%d") if bond.purchase_date else "",
            ]
        )
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    ws2 = wb.create_sheet("История сделок")
    headers2 = [
        "Название",
        "ISIN",
        "Кол-во",
        "Цена покупки (₽)",
        "Цена продажи (₽)",
        "Комиссия (₽)",
        "P&L (₽)",
        "P&L %",
        "Дата продажи",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="155724")
        cell.alignment = center
    for bond in sold:
        buy_p = float(bond.buy_price)
        sell_p = float(bond.sell_price) if bond.sell_price else buy_p
        comm = float(bond.broker_commission) if bond.broker_commission else 0.0
        pnl = round((sell_p - buy_p) * bond.amount - comm, 2)
        pnl_pct = round(pnl / (buy_p * bond.amount) * 100, 2) if buy_p else 0.0
        ws2.append(
            [
                bond.name,
                bond.isin,
                bond.amount,
                round(buy_p, 2),
                round(sell_p, 2),
                round(comm, 2),
                pnl,
                pnl_pct,
                bond.sell_date.strftime("%Y-%m-%d") if bond.sell_date else "",
            ]
        )
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = (
            max(len(str(cell.value or "")) for cell in col) + 4
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = (
        "attachment; filename=portfolio_report.xlsx"
    )
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response
